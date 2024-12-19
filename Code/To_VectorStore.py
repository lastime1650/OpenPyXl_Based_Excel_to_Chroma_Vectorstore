# 마이터 어택 정보(엑셀파일)를 벡터 스토어에 저장

from typing import Optional, Any, List, Dict

from langchain.document_loaders import WebBaseLoader, TextLoader, PyPDFLoader
from langchain.chains import RetrievalQA
from langchain.llms.ollama import Ollama
from langchain_community.embeddings import GPT4AllEmbeddings
from langchain_chroma import Chroma
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain.schema import Document
from langchain_community.vectorstores.utils import filter_complex_metadata # 복잡하고 양이 많은 메타데이터 사용시

import openpyxl

class MITRE_ATTACK_to_VectorStore():
    def __init__(self, EXEL_FILE_PATH:str, embedding_model:Optional[Any]=None):

        # 임베딩
        if embedding_model == None:
            embedding_model = GPT4AllEmbeddings()

        # 벡터스토어 매니저 생성
        self.VectorStore = Chroma(embedding_function=embedding_model,persist_directory=r"./VectorStore_data" )

        #--

        #sheets = openpyxl.load_workbook(EXEL_FILE_PATH)
        #self.Save_into_VectorStore(sheets)


        print( self.Query("adversaries may patch the authentication process on a domain controller to bypass the typical authentication mechanisms and enable access to accounts. malware may be used to inject false credentials", filter_metadata={"platform": 'Windows'}), True )
        return

    def Save_into_VectorStore(self, sheets:openpyxl.workbook.workbook.Workbook):
        for sheet in sheets:

            #

            if not ("techniques" == sheet.title or "software" == sheet.title): continue

            '''
                엑셀의 칼럼을 tuple로 변환하여 반환한다.
                ex) ('reference': 'A', 'citation': 'B', 'url': 'C')
            '''
            column_info = self.Column_info_to_dictionary(sheet)

            #  공통의 칼럼의 SHEET 찾기
            filter_columns = ["name", "description", "platforms"]
            if not (any(filter_column in column_info for filter_column in filter_columns)): continue

            # 이제 document 문장부분과 metadata를 분리해서 구현해야한다..
            documents_part = ["description"]
            metadata_part = ["name", "platforms"]

            '''
                난이도 높은 로직임.

                한 index에 추출한 ( 문서 및 여러 메타데이터들 )을 append 축적하는 구조임
                ex)
                [{
                    "document": ""; #  이녀석은 오로지 str
                    "metadata": [{"메타데이터1":"~~"}{"메타데이터2":"~~},,,] # List안에 Dictionary들이 위치하도록 구현한다.
                }]

            '''
            # 미리 문서관련 데이터를 사이즈를 잡아줘야한다.
            result_values: List[Optional[Dict]] = [{}] * (int(sheet.max_row) - 1)
            print(result_values)
            for content in column_info:
                if not (content in documents_part or content in metadata_part): continue

                # 문서 추출
                for document_part_name in documents_part:
                    if content in document_part_name:
                        # result_values[content] = []
                        print(content)
                        for i, row in enumerate(column_info[content]["sheet"]):

                            # 문서 문자 정리
                            def text_cleaner(text: str) -> str:
                                import re
                                text = text.replace("\n", "")

                                # '<' 과 '>' 포함된 문자를 제거
                                text = re.sub(r'<[^>]+>', '', text)

                                # '(' 과 ')' 이면서 그 안의 앞에 'http'가 포함되면 sub
                                text = re.sub(r'\(http[^)]+\)', '', text)

                                return text.lower()

                            # 문서는 "하나의 문자열로만 구성된다"
                            if "document" in result_values[i]:
                                result_values[i]["document"] += text_cleaner(row.value)
                            else:
                                result_values[i]["document"] = text_cleaner(row.value)

                # 메타데이터 추출
                for metadata_part_name in metadata_part:
                    if content in metadata_part_name:

                        # result_values[content] = []
                        print(content)
                        for i, row in enumerate(column_info[content]["sheet"]):
                            if "metadata" in result_values[i]:
                                if content in result_values[i]["metadata"]:
                                    result_values[i]["metadata"].append({
                                        content: row.value
                                    })
                                else:
                                    result_values[i]["metadata"].append({
                                        content: row.value
                                    })
                            else:
                                result_values[i] = {
                                    "metadata": [{
                                        content: row.value
                                    }]
                                }

                print("\n\n")

            # print(result_values)
            print(sheet)

            # VectorStore를 위한 "문서"생성
            Document_for_vectorstores = []
            for dict_data in result_values:
                if not ("document" in dict_data): print("스킵"); continue
                '''
                    ex) {"원하는 필터1": ~~~, "원하는 필터2": ~~~,,,} ( 1차원구조 ) 
                '''
                metadata_for_vectorstore = {}
                for metadata in dict_data["metadata"]:
                    metadata_for_vectorstore.update(metadata)
                print(metadata_for_vectorstore)
                Document_for_vectorstores.append(
                    Document(page_content=dict_data["document"], metadata=metadata_for_vectorstore))

            # 드디어 벡터스토어에 추가 !
            self.VectorStore.add_documents(filter_complex_metadata(Document_for_vectorstores))



    def Column_info_to_dictionary(self, sheet:openpyxl.workbook.workbook.Workbook)->Optional[dict]:

        # 열 수에 따라 '알파벳' 처리 (GPT)
        def column_index_to_letter(column_index):
            """
            열 인덱스(숫자)를 엑셀의 열 문자(A-ZZ)로 변환하는 함수입니다.

            Args:
              column_index: 변환할 열 인덱스 (1부터 시작).

            Returns:
              엑셀의 열 문자 (예: "A", "B", "Z", "AA", "AB", "ZZ").
            """

            if column_index < 1:
                raise ValueError("Column index must be greater than 0")

            result = ""
            while column_index > 0:
                remainder = (column_index - 1) % 26
                result = chr(65 + remainder) + result  # chr(65)는 'A'
                column_index = (column_index - 1) // 26

            return result

        def column_max_to_alphabet(column_max):
            """
            열 수를 엑셀의 열 문자(A-ZZ)로 변환하는 함수입니다.

            Args:
              column_max: 변환할 열 수 (1부터 시작).

            Returns:
              엑셀의 열 문자 (예: "A", "B", "Z", "AA", "AB", "ZZ").
            """
            return [column_index_to_letter(i) for i in range(1, column_max + 1)]

        #print(column_max_to_alphabet(sheet.max_column))
        column_letters = column_max_to_alphabet(sheet.max_column)

        def get_column_names(sheet, column_alphabets: List[str]) -> List[str]:
            output = []
            for column_letter in column_alphabets:
                column = sheet[column_letter][0]  # 0은 최상위 행(칼럼)
                output.append(column.value)
            return output

        columns_names = get_column_names(sheet, column_letters)

        # Result ( 최종은 DICT으로 변환 )
        result = {}
        for colum_info in zip(columns_names, column_letters):
            result[colum_info[0]] = {
                "column_letter": colum_info[1],
                "sheet": sheet[colum_info[1]][1:]
            }
        #print(result)

        #print("\n\n")
        return result


    def Add_Data(self, input_string:List[str], filter_metadata:List[dict] ):
        return self.VectorStore.add_texts(texts=input_string, metadatas=filter_metadata)

    def Query(self, input_string:str, filter_metadata:dict, is_verbose:bool = False ):
        if is_verbose:
            # 1에 가까우면 유사성이 높음
            return self.VectorStore.similarity_search_with_relevance_scores( query=input_string, k=3, filter=filter_metadata )
        else:
            for doc, score in self.VectorStore.similarity_search_with_relevance_scores( query=input_string, k=3, filter=filter_metadata ):
                print(f"문서:{doc}, 유사도점수:{score}\n")


EXCEL_FILE_PATH = "enterprise-attack-v16.1.xlsx"
MITRE_ATTACK_to_VectorStore(EXCEL_FILE_PATH)